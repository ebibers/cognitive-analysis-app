from tkinter import *
from tkinter import ttk
import tkinter.font as font
import random
import time
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

if not os.path.exists("scores.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "reaction game"
    wb.create_sheet("memory game")
    wb.create_sheet("target game")

    ws.append(["Date", "Score"])

    ws = wb["memory game"]
    ws.append(["Date", "Score"])

    ws = wb["target game"]
    ws.append(["Date", "Score"])

    wb.save("scores.xlsx")

def game_tab():
    game_frame.pack(fill = "both", expand = True)
    game_btn.configure(relief = SUNKEN)
    analysis_btn.configure(relief = RAISED)
    analysis_frame.forget()
    reaction_frame.forget()
    memory_frame.forget()
    target_frame.forget()
    reaction_frame_score.forget()
    memory_wrong.forget()
    target_score_frame.forget()
    analysis_dropdown.set("")
    reaction_text.configure(text = "Test your reactions.\nClick anywhere to start", bg = "white", fg = "black")
    reaction_frame.configure(bg = "white")
    globals()["in_progress"] = False

def analysis_tab():
    analysis_frame.pack(fill = "both", expand = True)
    game_btn.configure(relief = RAISED)
    analysis_btn.configure(relief = SUNKEN)
    game_frame.forget()
    reaction_frame.forget()
    memory_frame.forget()
    target_frame.forget()
    reaction_frame_score.forget()
    memory_wrong.forget()
    target_score_frame.forget()
    reaction_text.configure(text = "Test your reactions.\nClick anywhere to start", bg = "white", fg = "black")
    reaction_frame.configure(bg = "white")
    globals()["in_progress"] = False

def reaction_tab():
    reaction_frame.pack(fill = "both", expand = True)
    game_btn.configure(relief = RAISED)
    game_frame.forget()

def memory_tab():
    memory_frame.pack(fill = "both", expand = True)
    game_btn.configure(relief = RAISED)
    game_frame.forget()

def target_tab():
    target_frame.pack(fill = "both", expand = True)
    game_btn.configure(relief = RAISED)
    game_frame.forget()

def reaction_game_end(event):
    time_end = time.perf_counter()

    time_spent = time_end - time_start

    score = round(time_spent * 1000)

    wb = load_workbook("scores.xlsx")

    ws = wb["reaction game"]

    ws.append([datetime.now(), score])

    wb.save("scores.xlsx")

    reaction_frame_score.configure(bg = "white")
    reaction_text_score.configure(text = f"Your score is: {score}ms", bg = "white")

    reaction_frame_score.pack(fill = "both", expand = True)
    reaction_frame_continue.forget()

    game_btn.configure(state = "normal")
    analysis_btn.configure(state = "normal")

def reaction_game_continue(): 
    reaction_frame_continue.pack(fill = "both", expand = True)
    reaction_frame.forget()
    global time_start 
    time_start = time.perf_counter()

def reaction_game_start(event):
    if globals()["in_progress"] == True:
        reaction_frame.forget()
        reaction_frame.after_cancel(reaction_frame.after_id)
        game_btn.configure(state = "normal")
        analysis_btn.configure(state = "normal")
        reaction_frame_score.configure(bg = "yellow")
        reaction_text_score.configure(bg = "yellow", text = "Too soon!")
        reaction_frame_score.pack(fill = "both", expand = True)
        return
    
    globals()["in_progress"] = True

    game_btn.configure(state = "disabled")
    analysis_btn.configure(state = "disabled")

    reaction_text.configure(text = "Wait for green...", bg = "red", fg = "white")
    reaction_frame.configure(bg = "red")
    
    reaction_frame.after_id = reaction_frame.after(random.randint(1000, 5000), reaction_game_continue)

def reaction_game_reset():
    reaction_text.configure(text = "Test your reactions.\nClick anywhere to start", bg = "white", fg = "black")
    reaction_frame.configure(bg = "white")
    reaction_frame_score.forget()
    reaction_frame.pack(fill = "both", expand = True)
    globals()["in_progress"] = False

def get_input(input_number):
    input_start = memory_input_field.cget("text")

    if input_number == "Delete":
        input_start = input_start[:-1]
        full_input = input_start
    elif input_number == "Submit":
        memory_submit()
        return
    else:
        full_input = input_start + input_number

    memory_input_field.configure(text = full_input)

def memory_submit():
    user_answer = memory_input_field.cget("text")

    if str(user_answer) == str(memory_value):
        memory_input_frame.forget()
        memory_right.pack(fill = "both", expand = True)
        globals()["multiplier"] = globals()["multiplier"] * 10
        globals()["memory_score"] = globals()["memory_score"] + 1
        right_text.configure(text = f"Correct!\nCurrent score: {globals()["memory_score"]}")
    else:
        memory_input_frame.forget()
        memory_wrong.pack(fill = "both", expand = True)
        game_btn.configure(state = "normal")
        analysis_btn.configure(state = "normal")
        wrong_text.configure(text = f"Wrong answer!\nYour score is: {globals()["memory_score"]}")
        wrong_input.configure(text = f"Your answer:\n{user_answer}\nRight answer\n{memory_value}")

        wb = load_workbook("scores.xlsx")

        ws = wb["memory game"]

        ws.append([datetime.now(), globals()["memory_score"]])

        wb.save("scores.xlsx")

        globals()["in_progress"] = False

def memory_game_continue():
    memory_input_btn_0.configure(state = "normal")
    memory_input_btn_1.configure(state = "normal")
    memory_input_btn_2.configure(state = "normal")
    memory_input_btn_3.configure(state = "normal")
    memory_input_btn_4.configure(state = "normal")
    memory_input_btn_5.configure(state = "normal")
    memory_input_btn_6.configure(state = "normal")
    memory_input_btn_7.configure(state = "normal")
    memory_input_btn_8.configure(state = "normal")
    memory_input_btn_9.configure(state = "normal")
    memory_input_btn_delete.configure(state = "normal")
    memory_input_btn_submit.configure(state = "normal")
    memory_input_field.configure(text = "")

def memory_round():
    memory_right.forget()
    memory_input_frame.pack(fill = "both", expand = True)
    memory_input_btn_0.configure(state = "disabled")
    memory_input_btn_1.configure(state = "disabled")
    memory_input_btn_2.configure(state = "disabled")
    memory_input_btn_3.configure(state = "disabled")
    memory_input_btn_4.configure(state = "disabled")
    memory_input_btn_5.configure(state = "disabled")
    memory_input_btn_6.configure(state = "disabled")
    memory_input_btn_7.configure(state = "disabled")
    memory_input_btn_8.configure(state = "disabled")
    memory_input_btn_9.configure(state = "disabled")
    memory_input_btn_delete.configure(state = "disabled")
    memory_input_btn_submit.configure(state = "disabled")

    memory_input_field.configure(text = f"You have 5 seconds to memorize this number:\n{memory_value}")

    memory_input_field.after_id = memory_input_field.after(5000, memory_game_continue)

def memory_game_start():
    if globals()["in_progress"] == False:
        memory_wrong.forget()
        memory_frame.forget()
        memory_input_frame.pack(fill = "both", expand = True)

        game_btn.configure(state = "disabled")
        analysis_btn.configure(state = "disabled")

        globals()["in_progress"] = True

        global multiplier

        multiplier = 1

        global memory_score

        memory_score = 0

        global memory_value

        memory_value = random.randint(0, 9)

        memory_round()
    else:
        memory_value = random.randrange(1 * multiplier, 10 * multiplier)
        memory_round()

def target_game_end():
    time_end = time.perf_counter()
    time_spent = time_end - time_start

    game_btn.configure(state = "normal")
    analysis_btn.configure(state = "normal")
    target_area.forget()
    target_score_frame.pack(fill = "both", expand = True)
    target_score_text.configure(text = f"Your score is: {round(time_spent, 1)} seconds")

    wb = load_workbook("scores.xlsx")

    ws = wb["target game"]

    ws.append([datetime.now(), round(time_spent, 1)])

    wb.save("scores.xlsx")

def target_clicked(event):
    globals()["targets_destroyed"] += 1

    if globals()["targets_destroyed"] == 20:
        target_game_end()
    else:
        target_photo_label.place(relx = random.uniform(0, 0.9), rely = random.uniform(0, 0.85))

def target_game_start():
    target_score_frame.forget()
    target_frame.forget()
    target_area.pack(fill = "both", expand = True)
    game_btn.configure(state = "disabled")
    analysis_btn.configure(state = "disabled")

    global targets_destroyed

    targets_destroyed = 0

    target_photo_label.place(relx = random.uniform(0, 0.9), rely = random.uniform(0, 0.85))

    target_photo_label.bind("<Button-1>", target_clicked)

    globals()["time_start"] = time.perf_counter()

def show_analysis(game):
    ax.clear()
    if game == "Reaction game":
        scores = "scores.xlsx"

        df_reaction = pd.read_excel(scores, sheet_name = "reaction game")

        dates = df_reaction["Date"]
        scores = df_reaction["Score"]

        played = scores.count()
        games_played.configure(text = f"Games played: {played}")

        best = scores.min()
        best_score.configure(text = f"Best score: {best}ms")

        avg = scores.mean()
        avg_score.configure(text = f"Average score: {round(avg)}ms")

        if played >= 10:
            last_10 = scores.tail(10)
            avg_10 = last_10.mean()

            avg_score_10.configure(text = f"Average in last 10: {round(avg_10)}ms")
        else:
            avg_score_10.configure(text = "Average in last 10: -")

        canvas.get_tk_widget().pack()

        xvalues = []

        for x in range(len(dates)):
            xvalues.append(x)

        ax.plot(xvalues, scores)

        ax.set_title("Reaction Test Scores")
        ax.set_xlabel("Time")
        ax.set_ylabel("Reaction Time (ms)")

        ax.tick_params(labelbottom = False, bottom = False)

        canvas.draw()
    elif game == "Memory game":
        scores = "scores.xlsx"

        df_reaction = pd.read_excel(scores, sheet_name = "memory game")

        dates = df_reaction["Date"]
        scores = df_reaction["Score"]

        played = scores.count()
        games_played.configure(text = f"Games played: {played}")

        best = scores.max()
        best_score.configure(text = f"Best score: {best} digits")

        avg = scores.mean()
        avg_score.configure(text = f"Average score: {round(avg, 1)} digits")

        if played >= 10:
            last_10 = scores.tail(10)
            avg_10 = last_10.mean()

            avg_score_10.configure(text = f"Average in last 10: {round(avg_10, 1)} digits")
        else:
            avg_score_10.configure(text = "Average in last 10: -")

        canvas.get_tk_widget().pack()

        xvalues = []

        for x in range(len(dates)):
            xvalues.append(x)

        ax.plot(xvalues, scores)

        ax.set_title("Memory Test Scores")
        ax.set_xlabel("Time")
        ax.set_ylabel("Amount of digits memorized")

        ax.tick_params(labelbottom = False, bottom = False)

        canvas.draw()
    elif game == "Target game":
        scores = "scores.xlsx"

        df_reaction = pd.read_excel(scores, sheet_name = "target game")

        dates = df_reaction["Date"]
        scores = df_reaction["Score"]

        played = scores.count()
        games_played.configure(text = f"Games played: {played}")

        best = scores.min()
        best_score.configure(text = f"Best score: {best} sec")

        avg = scores.mean()
        avg_score.configure(text = f"Average score: {round(avg, 1)} sec")

        if played >= 10:
            last_10 = scores.tail(10)
            avg_10 = last_10.mean()

            avg_score_10.configure(text = f"Average in last 10: {round(avg_10, 1)} sec")
        else:
            avg_score_10.configure(text = "Average in last 10: -")

        canvas.get_tk_widget().pack()

        xvalues = []

        for x in range(len(dates)):
            xvalues.append(x)

        ax.plot(xvalues, scores)

        ax.set_title("Target Test Scores")
        ax.set_xlabel("Time")
        ax.set_ylabel("Time to destroy 20 targets (seconds)")

        ax.tick_params(labelbottom = False, bottom = False)

        canvas.draw()

def close_app():
    exit()

root = Tk()

root.geometry("960x540")
root.resizable(height = False, width = False)
root.title("Cognitive Ability Analysis Tool")

main_btn_frame = Frame(root)
main_btn_frame.columnconfigure(0, weight = 1)
main_btn_frame.columnconfigure(1, weight = 1)

analysis_photo = PhotoImage(file = r"icons\chart-histogram.png")
games_photo = PhotoImage(file = r"icons\computer.png")

btn_font = font.Font(family = "Calibri", weight = "bold", size = 16)

analysis_btn = Button(main_btn_frame, text = "  Analysis", command = analysis_tab, height = 60, font = btn_font, image = analysis_photo, compound = LEFT)
analysis_btn.grid(column = 1, row = 0, sticky = "we")

game_btn = Button(main_btn_frame, text = "  Games", command = game_tab, relief = SUNKEN, height = 60, font = btn_font, image = games_photo, compound = LEFT)
game_btn.grid(column = 0, row = 0, sticky = "we")

main_btn_frame.pack(fill = "x")

game_frame = Frame(root, bg = "white")
game_frame.columnconfigure(0, weight = 1)
game_frame.columnconfigure(1, weight = 1)
game_frame.columnconfigure(2, weight = 1)

reaction_photo = PhotoImage(file = r"icons\time-fast.png")
memory_photo = PhotoImage(file = r"icons\square-9.png")
target_photo = PhotoImage(file = r"icons\target.png")

reaction_btn = Button(game_frame, text = "Reaction Game", font = ("Calibri 16 bold"), image = reaction_photo, height = 350, width = 250, compound = TOP, command = reaction_tab)
reaction_btn.grid(column = 0, row = 0, pady = (50, 0), padx = (25, 0))

memory_btn = Button(game_frame, text = "Memory Game", font = ("Calibri 16 bold"), image = memory_photo, height = 350, width = 250, compound = TOP, command = memory_tab)
memory_btn.grid(column = 1, row = 0, pady = (50, 0))

target_btn = Button(game_frame, text = "Target Game", font = ("Calibri 16 bold"), image = target_photo, height = 350, width = 250, compound = TOP, command = target_tab)
target_btn.grid(column = 2, row = 0, pady = (50, 0), padx = (0, 25))

game_frame.pack(fill = "both", expand = True)

analysis_frame = Frame(root, bg = "white")

analysis_btn_frame = Frame(analysis_frame, borderwidth = 1, relief = "sunken")
analysis_btn_frame.pack(fill = "x", expand = True, anchor = "n")

analysis_btn_frame.columnconfigure(0, weight = 1)
analysis_btn_frame.columnconfigure(1, weight = 1)

analysis_dropdown_text = Label(analysis_btn_frame, text = "Select Game:")
analysis_dropdown_text.grid(column = 0, row = 0, sticky = "w", pady = (5, 5), padx = (15, 0))

analysis_dropdown = ttk.Combobox(analysis_btn_frame, values = ["Reaction game", "Memory game", "Target game"], state = "readonly")
analysis_dropdown.grid(column = 0, row = 0, pady = (5, 5), sticky = "w", padx = (90, 0))

analysis_dropdown.bind("<<ComboboxSelected>>",lambda e: analysis_btn_frame.focus())

analysis_analyze_btn = Button(analysis_btn_frame, text = "Show Statistics", width = 15, command = lambda : show_analysis(analysis_dropdown.get()))
analysis_analyze_btn.grid(column = 1, row = 0, sticky = "e", pady = (5, 5), padx = (0, 15))

analysis_statistics_frame = Frame(analysis_btn_frame, borderwidth = 1, relief = "sunken")
analysis_statistics_frame.grid(columnspan = 2, row = 1, sticky = "ew")

analysis_statistics_frame.columnconfigure(0, weight = 1)
analysis_statistics_frame.columnconfigure(1, weight = 1)
analysis_statistics_frame.columnconfigure(2, weight = 1)
analysis_statistics_frame.columnconfigure(3, weight = 1)

games_played = Label(analysis_statistics_frame, text = "Games played: -")
games_played.grid(column = 0, row = 0, pady = (5, 5))

avg_score = Label(analysis_statistics_frame, text = "Average score: -")
avg_score.grid(column = 1, row = 0)

best_score = Label(analysis_statistics_frame, text = "Best score: -")
best_score.grid(column = 2, row = 0)

avg_score_10 = Label(analysis_statistics_frame, text = "Average in last 10: -")
avg_score_10.grid(column = 3, row = 0)

analysis_graph_frame = Frame(analysis_frame, bg = "white")
analysis_graph_frame.pack(fill = "both", expand = True)

fig, ax = plt.subplots()

canvas = FigureCanvasTkAgg(fig, master = analysis_graph_frame)

reaction_frame = Frame(root, bg = "white")
reaction_frame.bind("<Button-1>", reaction_game_start)

global in_progress

in_progress = False

reaction_text = Label(reaction_frame, text = "Test your reactions.\nClick anywhere to start", bg = "white", font = ("Calibri 24 bold"))
reaction_text.place(relx = 0.5, rely = 0.5, anchor = CENTER)

reaction_text.bind("<Button-1>", reaction_game_start)

reaction_frame_continue = Frame(root, bg = "green")
reaction_frame_continue.bind("<Button-1>", reaction_game_end)

reaction_text_continue = Label(reaction_frame_continue, text = "Click!", bg = "green", font = ("Calibri 24 bold"), fg = "white")
reaction_text_continue.place(relx = 0.5, rely = 0.5, anchor = CENTER)

reaction_text_continue.bind("<Button-1>", reaction_game_end)

reaction_frame_score = Frame(root, bg = "white")

reaction_text_score = Label(reaction_frame_score, text = "...", bg = "white", font = ("Calibri 24 bold"))
reaction_text_score.place(relx = 0.5, rely = 0.5, anchor = CENTER)

reaction_score_button = Button(reaction_frame_score, text = "Play Again", command = reaction_game_reset)
reaction_score_button.place(relx = 0.5, rely = 0.6, anchor = CENTER)

memory_frame = Frame(root, bg = "white")

memory_text = Label(memory_frame, text = "Memory Game\nSee how many digits you can remember.", bg = "white", font = ("Calibri 24 bold"))
memory_text.place(relx = 0.5, rely = 0.5, anchor = CENTER)

memory_start = Button(memory_frame, text = "Start Game", command = memory_game_start)
memory_start.place(relx = 0.5, rely = 0.65, anchor = CENTER)

memory_input_frame = Frame(root, bg = "white")

memory_input_frame.columnconfigure(0, weight = 1)
memory_input_frame.columnconfigure(1, weight = 1)
memory_input_frame.columnconfigure(2, weight = 1)
memory_input_frame.columnconfigure(3, weight = 1)
memory_input_frame.columnconfigure(4, weight = 1)
memory_input_frame.columnconfigure(5, weight = 1)

memory_input_field = Label(memory_input_frame, text = "", height = 3, borderwidth = 1, relief = "solid", bg = "white", font = ("Calibri 32 bold"))
memory_input_field.grid(columnspan = 6, row = 0, sticky = "we")

memory_input_btn_0 = Button(memory_input_frame, text = "0", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("0"))
memory_input_btn_0.grid(column = 0, row = 1, pady = (35, 20), padx = (100, 0))

memory_input_btn_1 = Button(memory_input_frame, text = "1", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("1"))
memory_input_btn_1.grid(column = 1, row = 1, pady = (35, 20))

memory_input_btn_2 = Button(memory_input_frame, text = "2", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("2"))
memory_input_btn_2.grid(column = 2, row = 1, pady = (35, 20))

memory_input_btn_3 = Button(memory_input_frame, text = "3", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("3"))
memory_input_btn_3.grid(column = 3, row = 1, pady = (35, 20))

memory_input_btn_4 = Button(memory_input_frame, text = "4", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("4"))
memory_input_btn_4.grid(column = 4, row = 1, pady = (35, 20))

memory_input_btn_5 = Button(memory_input_frame, text = "5", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("5"))
memory_input_btn_5.grid(column = 0, row = 2, padx = (100, 0))

memory_input_btn_6 = Button(memory_input_frame, text = "6", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("6"))
memory_input_btn_6.grid(column = 1, row = 2)

memory_input_btn_7 = Button(memory_input_frame, text = "7", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("7"))
memory_input_btn_7.grid(column = 2, row = 2)

memory_input_btn_8 = Button(memory_input_frame, text = "8", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("8"))
memory_input_btn_8.grid(column = 3, row = 2)

memory_input_btn_9 = Button(memory_input_frame, text = "9", font = ("Calibri 24 bold"), width = 5, height = 2, command = lambda : get_input("9"))
memory_input_btn_9.grid(column = 4, row = 2)

memory_input_btn_delete = Button(memory_input_frame, text = "Delete", font = ("Calibri 24 bold"), width = 6, height = 2, command = lambda : get_input("Delete"))
memory_input_btn_delete.grid(column = 5, row = 1, padx = (0, 100), pady = (15, 0))

memory_input_btn_submit = Button(memory_input_frame, text = "Submit", font = ("Calibri 24 bold"), width = 6, height = 2, command = lambda : get_input("Submit"))
memory_input_btn_submit.grid(column = 5, row = 2, padx = (0, 100))

memory_wrong = Frame(root, bg = "red")

wrong_text = Label(memory_wrong, text = "Wrong answer", font = ("Calibri 24 bold"), bg = "red", fg = "white")
wrong_text.place(relx = 0.5, rely = 0.25, anchor = CENTER)

wrong_input = Label(memory_wrong, text = "...", font = ("Calibri 18 bold"), bg = "red", fg = "white")
wrong_input.place(relx = 0.5, rely = 0.5, anchor = CENTER)

wrong_btn = Button(memory_wrong, text = "Try again", command = memory_game_start)
wrong_btn.place(relx = 0.5, rely = 0.7, anchor = CENTER)

memory_right = Frame(root, bg = "green")

right_text = Label(memory_right, text = "...", font = ("Calibri 24 bold"), bg = "green", fg = "white")
right_text.place(relx = 0.5, rely = 0.5, anchor = CENTER)

right_btn = Button(memory_right, text = "Continue", command = memory_game_start)
right_btn.place(relx = 0.5, rely = 0.65, anchor = CENTER)

target_frame = Frame(root, bg = "white")

target_text = Label(target_frame, text = "Target Game\nSee how fast you can click 20 targets.", font = ("Calibri 24 bold"), bg = "white")
target_text.place(relx = 0.5, rely = 0.5, anchor = CENTER)

target_start_btn = Button(target_frame, text = "Start Game", command = target_game_start)
target_start_btn.place(relx = 0.5, rely = 0.65, anchor = CENTER)

target_area = Frame(root, bg = "white")

target_small_photo = PhotoImage(file = r"icons\target_small.png")

target_photo_label = Label(target_area, image = target_small_photo, bg = "white")
target_photo_label.image = target_small_photo

target_score_frame = Frame(root, bg = "white")

target_score_text = Label(target_score_frame, text = "...", bg = "white", font = ("Calibri 24 bold"))
target_score_text.place(relx = 0.5, rely = 0.5, anchor = CENTER)

target_score_btn = Button(target_score_frame, text = "Try again", command = target_game_start)
target_score_btn.place(relx = 0.5, rely = 0.6, anchor = CENTER)

root.protocol("WM_DELETE_WINDOW", close_app)
root.mainloop()